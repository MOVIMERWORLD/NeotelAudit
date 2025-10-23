#!/usr/bin/env python3
"""
Convertidor de snapshots JSON a Excel Profesional
Genera un archivo Excel corporativo con estilo minimalista
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importar configuración
from config_audit import ConfigAudit

class ExcelProfessionalConverter:
    """Convertidor profesional de JSON a Excel con estilo corporativo"""
    
    def __init__(self):
        """Inicializar convertidor con paleta corporativa"""
        self.workbook = Workbook()
        # Eliminar hoja por defecto
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Paleta corporativa minimalista (igual al dashboard)
        self.colors = {
            'primary': '2C3E50',           # Azul oscuro corporativo
            'accent': '3498DB',            # Azul moderado
            'success': '27AE60',           # Verde
            'warning': 'F39C12',           # Naranja
            'danger': 'E74C3C',            # Rojo
            'background': 'F8F9FA',        # Gris muy claro
            'surface': 'FFFFFF',           # Blanco
            'border': 'DEE2E6',            # Gris claro
            'text_secondary': '7F8C8D',    # Gris medio
            
            # Colores suaves para datos
            'active_light': 'D4EDDA',      # Verde claro
            'paused_light': 'F8D7DA',      # Rojo claro
            'mixed_light': 'FFF3CD',       # Amarillo claro
        }
    
    def apply_title_style(self, cell):
        """Estilo para títulos principales"""
        cell.font = Font(bold=True, size=16, color=self.colors['primary'])
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def apply_header_style(self, cell):
        """Estilo para encabezados de tabla"""
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=self.colors['primary'], 
                               end_color=self.colors['primary'], 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        cell.border = border
    
    def apply_data_cell_style(self, cell, bg_color=None, bold=False, align='left'):
        """Estilo para celdas de datos"""
        cell.font = Font(size=10, bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, 
                                   end_color=bg_color, 
                                   fill_type='solid')
        
        border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
        cell.border = border
    
    def auto_adjust_columns(self, worksheet):
        """Ajustar automáticamente el ancho de columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Ancho ajustado con límites
            adjusted_width = min(max(max_length + 3, 12), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_extensiones_sheet(self, extensiones):
        """Crear pestaña de extensiones con estilo profesional"""
        ws = self.workbook.create_sheet("📞 Extensiones")
        
        # Título
        ws['A1'] = '📞 Configuración de Extensiones'
        self.apply_title_style(ws['A1'])
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Espacio
        ws.row_dimensions[2].height = 10
        
        # Encabezados (con nueva columna)
        headers = ['Extensión', 'Nombre', 'Grupo', 'Número Saliente', 'Estado Colas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.apply_header_style(cell)
        
        # Datos
        for row_idx, ext in enumerate(extensiones, 4):
            estado = ext.get('estado_colas', 'sin_colas')
            
            # Determinar color según estado
            estado_config = {
                'todas_activas': ('✓ Activas', self.colors['active_light']),
                'todas_pausadas': ('⏸ Pausadas', self.colors['paused_light']),
                'mixto': ('⚡ Mixto', self.colors['mixed_light']),
                'sin_colas': ('○ Sin Colas', self.colors['surface'])
            }
            
            estado_label, bg_color = estado_config.get(estado, ('', self.colors['surface']))
            
            # Escribir datos (con nuevo campo)
            data = [
                ext.get('extension', ''),
                ext.get('nombre', ''),
                ext.get('agente_asignado', '-'),  # ⭐ NUEVO CAMPO
                ext.get('numero_saliente', '-'),
                estado_label
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                is_bold = (col_idx == 1)  # Bold en extensión
                self.apply_data_cell_style(cell, bg_color=bg_color, bold=is_bold)
        
        # Congelar filas de encabezado
        ws.freeze_panes = 'A4'
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A3:E{len(extensiones) + 3}"  # ⭐ Cambiar E por F
        
        self.auto_adjust_columns(ws)

    def create_dids_sheet(self, dids):
        """Crear pestaña de DIDs con estilo profesional"""
        ws = self.workbook.create_sheet("📱 DIDs")
        
        # Título
        ws['A1'] = '📱 Configuración de DIDs'
        self.apply_title_style(ws['A1'])
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Espacio
        ws.row_dimensions[2].height = 10
        
        # Encabezados
        headers = ['Número', 'Locución', 'Acción 1', 'Acción 2', 'Acción 3']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.apply_header_style(cell)
        
        # Datos
        for row_idx, did in enumerate(dids, 4):
            data = [
                did.get('numero', ''),
                did.get('locucion', 'N/D'),
                did.get('accion1', 'N/D'),
                did.get('accion2', 'N/D'),
                did.get('accion3', 'N/D')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                is_bold = (col_idx == 1)  # Bold en número
                self.apply_data_cell_style(cell, bold=is_bold)
        
        # Congelar filas de encabezado
        ws.freeze_panes = 'A4'
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A3:E{len(dids) + 3}"
        
        self.auto_adjust_columns(ws)
    
    def create_colas_sheet(self, colas):
        """Crear pestaña de colas con resumen"""
        ws = self.workbook.create_sheet("📋 Colas - Resumen")
        
        # Título
        ws['A1'] = '📋 Resumen de Colas'
        self.apply_title_style(ws['A1'])
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Espacio
        ws.row_dimensions[2].height = 10
        
        # Encabezados
        headers = ['Cola', 'Total Miembros', '▶ Activos', '⏸ Pausados']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.apply_header_style(cell)
        
        # Datos
        for row_idx, cola in enumerate(colas, 4):
            miembros = cola.get('miembros', [])
            activos = sum(1 for m in miembros if m.get('estado') == 'activo')
            pausados = sum(1 for m in miembros if m.get('estado') == 'pausado')
            
            data = [
                cola.get('nombre', ''),
                len(miembros),
                activos,
                pausados
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                is_bold = (col_idx == 1)
                align = 'center' if col_idx > 1 else 'left'
                self.apply_data_cell_style(cell, bold=is_bold, align=align)
        
        # Congelar filas de encabezado
        ws.freeze_panes = 'A4'
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A3:D{len(colas) + 3}"
        
        self.auto_adjust_columns(ws)
    
    def create_miembros_sheet(self, colas, extensiones):
        """Crear pestaña detallada de miembros con nombres (mejorado)"""
        ws = self.workbook.create_sheet("📋 Colas - Detalle")
        
        # Título
        ws['A1'] = '📋 Miembros de Colas (Detallado)'
        self.apply_title_style(ws['A1'])
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        # Espacio
        ws.row_dimensions[2].height = 10
        
        # Encabezados
        headers = ['Cola', 'Nombre Usuario', 'Extensión', 'Estado', 'Información']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self.apply_header_style(cell)
        
        # Crear diccionario de extensiones con NORMALIZACIÓN
        # Normalizar: eliminar espacios, guiones, etc para hacer match robusto
        ext_dict = {}
        ext_dict_raw = {}  # Diccionario con códigos originales
        
        for ext in extensiones:
            codigo_original = ext.get('extension', '').strip()
            nombre = ext.get('nombre', '').strip()
            
            # Guardar con código original
            ext_dict_raw[codigo_original] = nombre
            
            # También guardar versiones normalizadas para búsqueda flexible
            codigo_normalizado = codigo_original.replace('-', '').replace(' ', '').upper()
            if codigo_normalizado:
                ext_dict[codigo_normalizado] = nombre
        
        # Debug: mostrar algunos mapeos
        print(f"\n🔍 Debug - Primeras 5 extensiones en diccionario:")
        for i, (codigo, nombre) in enumerate(list(ext_dict_raw.items())[:5]):
            print(f"   {codigo} -> {nombre}")
        
        # Datos
        row_idx = 4
        extensiones_no_encontradas = set()
        
        for cola in colas:
            cola_nombre = cola.get('nombre', '')
            
            for miembro in cola.get('miembros', []):
                estado = miembro.get('estado', 'desconocido')
                extension_code = miembro.get('extension', '').strip()
                
                # BÚSQUEDA MEJORADA DEL NOMBRE
                nombre_usuario = None
                
                # 1. Intentar match exacto
                if extension_code in ext_dict_raw:
                    nombre_usuario = ext_dict_raw[extension_code]
                
                # 2. Intentar match normalizado
                if not nombre_usuario:
                    codigo_normalizado = extension_code.replace('-', '').replace(' ', '').upper()
                    if codigo_normalizado in ext_dict:
                        nombre_usuario = ext_dict[codigo_normalizado]
                
                # 3. Intentar buscar por subcadena (últimos dígitos)
                if not nombre_usuario:
                    # Extraer solo números del código
                    numeros_codigo = ''.join(filter(str.isdigit, extension_code))
                    
                    for codigo_ext, nombre_ext in ext_dict_raw.items():
                        numeros_ext = ''.join(filter(str.isdigit, codigo_ext))
                        if numeros_codigo and numeros_ext and numeros_codigo in numeros_ext:
                            nombre_usuario = nombre_ext
                            break
                
                # 4. Si aún no se encuentra, usar código como fallback
                if not nombre_usuario or nombre_usuario == '':
                    nombre_usuario = f"⚠️ {extension_code}"
                    extensiones_no_encontradas.add(extension_code)
                
                # Color y label según estado
                if estado == 'activo':
                    bg_color = self.colors['active_light']
                    estado_label = '▶ Activo'
                elif estado == 'pausado':
                    bg_color = self.colors['paused_light']
                    estado_label = '⏸ Pausado'
                else:
                    bg_color = self.colors['surface']
                    estado_label = '? Desconocido'
                
                # Extraer info limpia
                texto = miembro.get('texto', '')
                info = texto.split('Extensión:')[0].strip() if 'Extensión:' in texto else f"Ext: {extension_code}"
                
                data = [
                    cola_nombre,
                    nombre_usuario,
                    extension_code,
                    estado_label,
                    info
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    is_bold = (col_idx == 2)  # Bold en nombre
                    self.apply_data_cell_style(cell, bg_color=bg_color, bold=is_bold)
                
                row_idx += 1
        
        # Mostrar advertencia si hay extensiones no encontradas
        if extensiones_no_encontradas:
            print(f"\n⚠️  ADVERTENCIA: {len(extensiones_no_encontradas)} extensiones no encontradas:")
            for ext_code in sorted(list(extensiones_no_encontradas))[:10]:
                print(f"   - {ext_code}")
            if len(extensiones_no_encontradas) > 10:
                print(f"   ... y {len(extensiones_no_encontradas) - 10} más")
        
        # Congelar filas de encabezado
        ws.freeze_panes = 'A4'
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A3:E{row_idx - 1}"
        
        self.auto_adjust_columns(ws)

    def convert_json_to_excel(self, json_file, output_file=None):
        """Convertir archivo JSON a Excel profesional"""
        try:
            print(f"📖 Leyendo archivo JSON: {json_file}")
            
            # Leer JSON
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Extraer configuración
            if 'config' in data:
                config_data = data['config']
            else:
                config_data = data
            
            print("🎨 Generando Excel profesional...")
            
            extensiones = config_data.get('extensiones', [])
            dids = config_data.get('dids', [])
            colas = config_data.get('colas', [])
            
            # Crear hojas en orden (SIN DASHBOARD)
            self.create_extensiones_sheet(extensiones)
            self.create_dids_sheet(dids)
            self.create_colas_sheet(colas)
            self.create_miembros_sheet(colas, extensiones)
            
            # Generar nombre de archivo si no se proporciona
            if output_file is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = ConfigAudit.REPORTS_DIR / f"neotel_config_{timestamp}.xlsx"
            
            # Guardar
            self.workbook.save(output_file)
            
            print()
            print("=" * 80)
            print("✅ EXCEL PROFESIONAL GENERADO")
            print("=" * 80)
            print(f"\n📊 Archivo: {output_file}")
            print(f"📋 Pestañas creadas: {len(self.workbook.sheetnames)}")
            print(f"   • {len(extensiones)} Extensiones")
            print(f"   • {len(dids)} DIDs")
            print(f"   • {len(colas)} Colas (resumen y detalle)")
            
            return output_file
            
        except Exception as e:
            print(f"❌ Error convirtiendo JSON a Excel: {str(e)}")
            raise

def convert_latest_snapshot():
    """Convertir el snapshot más reciente a Excel"""
    try:
        # Buscar el snapshot más reciente
        snapshots = sorted(ConfigAudit.SNAPSHOTS_DIR.glob("*.json"), reverse=True)
        
        if not snapshots:
            print("❌ No se encontraron snapshots")
            return
        
        latest_snapshot = snapshots[0]
        print(f"📁 Snapshot más reciente: {latest_snapshot.name}")
        print()
        
        # Convertir
        converter = ExcelProfessionalConverter()
        excel_file = converter.convert_json_to_excel(latest_snapshot)
        
        print(f"\n💡 Abre el archivo Excel para visualizar los datos")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        sys.exit(1)


def convert_specific_snapshot(date_str):
    """Convertir un snapshot específico por fecha"""
    try:
        snapshot_file = ConfigAudit.SNAPSHOTS_DIR / f"{date_str}_snapshot.json"
        
        if not snapshot_file.exists():
            print(f"❌ No se encontró snapshot para la fecha {date_str}")
            return
        
        print(f"📁 Snapshot encontrado: {snapshot_file.name}")
        print()
        
        # Convertir
        converter = ExcelProfessionalConverter()
        excel_file = converter.convert_json_to_excel(snapshot_file)
        
        print(f"\n💡 Abre el archivo Excel para visualizar los datos")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        sys.exit(1)


def main():
    """Función principal"""
    print("=" * 80)
    print("📊 CONVERTIDOR JSON A EXCEL PROFESIONAL - NEOTEL")
    print("=" * 80)
    print()
    
    if len(sys.argv) > 1:
        # Convertir snapshot específico
        date_str = sys.argv[1]
        print(f"Convirtiendo snapshot de fecha: {date_str}")
        convert_specific_snapshot(date_str)
    else:
        # Convertir el más reciente
        print("Convirtiendo snapshot más reciente...")
        convert_latest_snapshot()


if __name__ == "__main__":
    main()